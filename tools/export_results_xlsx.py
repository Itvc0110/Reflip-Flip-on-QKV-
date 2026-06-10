"""Export experiment results to a Smart-Flip-style XLSX workbook.

The exporter reads the experiment ledger, lm-eval JSON outputs, and optional
perplexity JSON files produced by `compare_awq_slicing.py --save-json`.

Example:
    python tools/export_results_xlsx.py \
        --ledger results/experiment_ledger.jsonl \
        --lm-eval-dir results/lm_eval \
        --output results/smart_flip_results.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


TASK_COLUMNS = [
    "arc-c",
    "arc-e",
    "boolq",
    "hellaswag",
    "lambada",
    "openbookqa",
    "piqa",
    "rte",
    "winogrande",
]

SUMMARY_COLUMNS = [
    "model",
    "method",
    "bits",
    "group_size",
    "wiki_ppl",
    "c4_ppl",
    "ppl_avg",
    *TASK_COLUMNS,
    "avg",
    "source",
    "timestamp",
    "notes",
]

PPL_COLUMNS = [
    "model",
    "method",
    "dataset",
    "perplexity",
    "total_tokens",
    "n_samples",
    "model_path",
    "source",
    "timestamp",
]

LM_EVAL_COLUMNS = [
    "model",
    "method",
    "task",
    "workbook_task",
    "metric",
    "score",
    "source",
    "timestamp",
]

QKV_COLUMNS = [
    "result_file",
    "num_heads",
    "nearest_mean_abs_attention_error",
    "flip_mean_abs_attention_error",
    "reflip_mean_abs_attention_error",
    "nearest_to_flip_mean_improvement_pct",
    "nearest_to_reflip_mean_improvement_pct",
    "flip_to_reflip_mean_improvement_pct",
    "source",
    "timestamp",
]

MISSING_COLUMNS = ["model", "method", "missing_metric", "source"]

TASK_ALIASES = {
    "arc_challenge": "arc-c",
    "arc-c": "arc-c",
    "arc_easy": "arc-e",
    "arc-e": "arc-e",
    "boolq": "boolq",
    "hellaswag": "hellaswag",
    "lambada_openai": "lambada",
    "lambada": "lambada",
    "openbookqa": "openbookqa",
    "piqa": "piqa",
    "rte": "rte",
    "winogrande": "winogrande",
}

PREFERRED_LM_METRICS = [
    "acc_norm,none",
    "acc,none",
    "exact_match,none",
    "perplexity,none",
    "word_perplexity,none",
    "byte_perplexity,none",
]


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows = []
    for line_no, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            value = json.loads(line)
        except json.JSONDecodeError as exc:
            print(f"Skipping invalid ledger row {line_no} in {path}: {exc}")
            continue
        if isinstance(value, dict):
            rows.append(value)
    return rows


def find_json_files(path: Path) -> list[Path]:
    if path.is_file():
        return [path]
    if not path.exists():
        return []
    return sorted(p for p in path.rglob("*.json") if p.is_file())


def unique_join(values: list[str]) -> str:
    seen = []
    for value in values:
        if value and value not in seen:
            seen.append(value)
    return "; ".join(seen)


def infer_method(text: str | None) -> str:
    if not text:
        return ""
    lowered = text.replace("\\", "/").lower()
    if "awq_gqa" in lowered or "gqa_reflip" in lowered:
        return "awq_gqa_reflip"
    if "awq_js" in lowered or "james" in lowered:
        return "james_stein_heuristic_awq"
    if "awq_dh" in lowered or "dynamic" in lowered:
        return "dynamic_heuristic_awq"
    if "awq_standard" in lowered or "standard_awq" in lowered:
        return "standard_awq"
    if "rtn" in lowered or "round-to-nearest" in lowered:
        return "rtn_w4g128"
    if "gptq" in lowered:
        return "gptq"
    if "adaround" in lowered:
        return "adaround"
    if "qkv" in lowered or "standalone" in lowered:
        return "standalone_qkv"
    return Path(lowered).name or text


def infer_model(text: str | None) -> str:
    if not text:
        return ""
    lowered = text.replace("\\", "/").lower()
    if "llama-3.1" in lowered or "llama3.1" in lowered or "llama_3p1" in lowered:
        return "Llama-3.1-8B"
    if "llama-3" in lowered or "llama3" in lowered:
        return "Llama-3-8B"
    if "mistral" in lowered:
        return "Mistral-7B"
    if "qwen" in lowered:
        return "Qwen2.5-7B"
    if "minicpm" in lowered:
        return "MiniCPM-2B"
    return ""


def infer_bits(text: str | None) -> int | None:
    if not text:
        return None
    lowered = text.lower()
    if re.search(r"(^|[^0-9])3\s*bit", lowered) or "w3" in lowered or "_b3" in lowered:
        return 3
    if re.search(r"(^|[^0-9])4\s*bit", lowered) or "w4" in lowered or "_b4" in lowered:
        return 4
    if any(token in lowered for token in ["awq", "rtn", "gqa"]):
        return 4
    return None


def infer_group_size(text: str | None) -> int | None:
    if not text:
        return None
    match = re.search(r"g(\d+)", text.lower())
    if match:
        return int(match.group(1))
    if "group_size" in text.lower() or "w4" in text.lower() or "awq" in text.lower():
        return 128
    return None


def choose_lm_metric(task_result: dict[str, Any]) -> tuple[str, float] | None:
    for key in PREFERRED_LM_METRICS:
        value = task_result.get(key)
        if isinstance(value, (int, float)):
            return key, float(value)
    for key, value in task_result.items():
        if key.endswith("_stderr") or key.endswith(",stderr"):
            continue
        if isinstance(value, (int, float)):
            return key, float(value)
    return None


def metric_average(row: dict[str, Any], columns: list[str]) -> float | None:
    values = [row.get(col) for col in columns]
    numeric = [float(value) for value in values if isinstance(value, (int, float))]
    if not numeric:
        return None
    return sum(numeric) / len(numeric)


def add_source(summary: dict[tuple[str, str], dict[str, Any]], key: tuple[str, str], source: str) -> None:
    row = summary.setdefault(key, {"model": key[0], "method": key[1], "_sources": []})
    row.setdefault("_sources", []).append(source)


def update_timestamp(row: dict[str, Any], timestamp: str | None) -> None:
    if timestamp and not row.get("timestamp"):
        row["timestamp"] = timestamp


def parse_ppl_json(path: Path) -> list[dict[str, Any]]:
    data = read_json(path)
    timestamp = data.get("timestamp")
    n_samples = data.get("n_samples")
    model_paths = {
        "Heuristic AWQ": data.get("heuristic_path"),
        "Standard AWQ": data.get("standard_path"),
    }

    rows = []
    raw_results = data.get("raw_results", {})
    if isinstance(raw_results, dict):
        for dataset_name, models in raw_results.items():
            if not isinstance(models, dict):
                continue
            for model_label, result in models.items():
                if not isinstance(result, dict):
                    continue
                model_path = model_paths.get(model_label)
                method_hint = model_path or model_label
                row = {
                    "model": infer_model(model_path) or infer_model(model_label),
                    "method": infer_method(method_hint),
                    "dataset": dataset_name,
                    "perplexity": result.get("perplexity"),
                    "total_tokens": result.get("total_tokens"),
                    "n_samples": n_samples,
                    "model_path": model_path,
                    "source": str(path),
                    "timestamp": timestamp,
                }
                rows.append(row)
    return rows


def parse_lm_eval_json(path: Path) -> list[dict[str, Any]]:
    data = read_json(path)
    results = data.get("results")
    if not isinstance(results, dict):
        return []

    config = data.get("config") if isinstance(data.get("config"), dict) else {}
    source_text = " ".join(
        str(part)
        for part in [
            path,
            config.get("model"),
            config.get("model_args"),
            data.get("model_args"),
        ]
        if part
    )
    rows = []
    for task_name, task_result in sorted(results.items()):
        if not isinstance(task_result, dict):
            continue
        chosen = choose_lm_metric(task_result)
        if chosen is None:
            continue
        metric, score = chosen
        workbook_task = TASK_ALIASES.get(task_name, task_name)
        rows.append(
            {
                "model": infer_model(source_text),
                "method": infer_method(source_text),
                "task": task_name,
                "workbook_task": workbook_task,
                "metric": metric,
                "score": score,
                "source": str(path),
                "timestamp": data.get("date") or data.get("timestamp"),
            }
        )
    return rows


def parse_lm_eval_ledger(row: dict[str, Any]) -> list[dict[str, Any]]:
    scores = row.get("workbook_scores")
    if not isinstance(scores, dict):
        return []
    source = row.get("file") or row.get("label") or "results/experiment_ledger.jsonl"
    source_text = str(source)
    rows = []
    for workbook_task, score in scores.items():
        if not isinstance(score, (int, float)):
            continue
        rows.append(
            {
                "model": infer_model(source_text),
                "method": infer_method(source_text),
                "task": workbook_task,
                "workbook_task": TASK_ALIASES.get(workbook_task, workbook_task),
                "metric": "ledger_score",
                "score": float(score),
                "source": source_text,
                "timestamp": row.get("timestamp"),
            }
        )
    return rows


def parse_qkv_ledger(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    qkv_rows = []
    for row in rows:
        if row.get("phase") != "standalone_qkv_summary":
            continue
        metrics = row.get("metrics")
        if not isinstance(metrics, dict):
            continue
        qkv_row = {
            "result_file": row.get("result_file"),
            "num_heads": row.get("num_heads"),
            "source": row.get("script", "tools/summarize_qkv_results.py"),
            "timestamp": row.get("timestamp"),
        }
        qkv_row.update(metrics)
        qkv_rows.append(qkv_row)
    return qkv_rows


def build_summary_rows(
    ledger_rows: list[dict[str, Any]],
    ppl_rows: list[dict[str, Any]],
    lm_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    summary: dict[tuple[str, str], dict[str, Any]] = {}

    for row in ledger_rows:
        method = row.get("method") or infer_method(row.get("output_dir") or row.get("script") or row.get("phase"))
        model = infer_model(row.get("model_path") or row.get("output_dir") or row.get("file") or row.get("label"))
        if not method or not model:
            continue
        key = (model, method)
        add_source(summary, key, "results/experiment_ledger.jsonl")
        target = summary[key]
        target.setdefault("bits", row.get("bits") or infer_bits(str(row)))
        target.setdefault("group_size", row.get("group_size") or infer_group_size(str(row)))
        target.setdefault("notes", row.get("notes"))
        update_timestamp(target, row.get("timestamp"))

    for row in ppl_rows:
        model = row.get("model") or infer_model(row.get("model_path"))
        method = row.get("method") or infer_method(row.get("model_path"))
        if not model or not method:
            continue
        key = (model, method)
        add_source(summary, key, row.get("source", "ppl_json"))
        target = summary[key]
        target.setdefault("bits", infer_bits(row.get("model_path") or method))
        target.setdefault("group_size", infer_group_size(row.get("model_path") or method))
        update_timestamp(target, row.get("timestamp"))
        dataset = str(row.get("dataset", "")).lower()
        if dataset in {"wikitext-2", "wikitext2", "wiki"}:
            target["wiki_ppl"] = row.get("perplexity")
        elif dataset == "c4":
            target["c4_ppl"] = row.get("perplexity")

    for row in lm_rows:
        model = row.get("model") or infer_model(row.get("source"))
        method = row.get("method") or infer_method(row.get("source"))
        task = row.get("workbook_task")
        if not model or not method or task not in TASK_COLUMNS:
            continue
        key = (model, method)
        add_source(summary, key, row.get("source", "lm_eval"))
        target = summary[key]
        target.setdefault("bits", infer_bits(row.get("source") or method))
        target.setdefault("group_size", infer_group_size(row.get("source") or method))
        target[task] = row.get("score")
        update_timestamp(target, row.get("timestamp"))

    final_rows = []
    for row in summary.values():
        row["ppl_avg"] = metric_average(row, ["wiki_ppl", "c4_ppl"])
        row["avg"] = metric_average(row, TASK_COLUMNS)
        row["source"] = unique_join(row.pop("_sources", []))
        final_rows.append({column: row.get(column) for column in SUMMARY_COLUMNS})

    return sorted(final_rows, key=lambda item: (str(item.get("model")), str(item.get("method"))))


def build_missing_rows(summary_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    required = ["wiki_ppl", "c4_ppl", *TASK_COLUMNS]
    rows = []
    for row in summary_rows:
        if not row.get("model") or not row.get("method"):
            continue
        for metric in required:
            if row.get(metric) is None:
                rows.append(
                    {
                        "model": row.get("model"),
                        "method": row.get("method"),
                        "missing_metric": metric,
                        "source": row.get("source"),
                    }
                )
    return rows


def write_sheet(workbook: Workbook, title: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
    if title in workbook.sheetnames:
        del workbook[title]
    sheet = workbook.create_sheet(title)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for col_idx, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=column)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, column in enumerate(columns, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=row.get(column))

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for col_idx, column in enumerate(columns, start=1):
        max_len = len(column)
        for row_idx in range(2, min(sheet.max_row, 200) + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def load_or_create_workbook(template: Path | None) -> Workbook:
    if template and template.exists():
        workbook = load_workbook(template)
    else:
        workbook = Workbook()

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
        del workbook["Sheet"]
    return workbook


def main() -> int:
    parser = argparse.ArgumentParser(description="Export Smart-Flip-style experiment workbook")
    parser.add_argument("--ledger", default="results/experiment_ledger.jsonl")
    parser.add_argument("--lm-eval-dir", default="results/lm_eval")
    parser.add_argument("--ppl-dir", default="results/ppl")
    parser.add_argument("--output", default="results/smart_flip_results.xlsx")
    parser.add_argument("--template", default="", help="Optional workbook template to copy sheets from")
    args = parser.parse_args()

    ledger_path = Path(args.ledger)
    lm_eval_dir = Path(args.lm_eval_dir)
    ppl_dir = Path(args.ppl_dir)
    output_path = Path(args.output)
    template_path = Path(args.template) if args.template else None

    ledger_rows = read_jsonl(ledger_path)
    qkv_rows = parse_qkv_ledger(ledger_rows)

    ppl_rows = []
    for path in find_json_files(ppl_dir):
        try:
            ppl_rows.extend(parse_ppl_json(path))
        except Exception as exc:  # noqa: BLE001 - keep exporting other files.
            print(f"Skipping PPL JSON {path}: {exc}")

    lm_rows = []
    seen_lm_files = set()
    for path in find_json_files(lm_eval_dir):
        try:
            parsed = parse_lm_eval_json(path)
        except Exception as exc:  # noqa: BLE001 - keep exporting other files.
            print(f"Skipping lm-eval JSON {path}: {exc}")
            continue
        if parsed:
            seen_lm_files.add(str(path))
            lm_rows.extend(parsed)

    for row in ledger_rows:
        source_file = str(row.get("file", ""))
        if row.get("phase") == "lm_eval_summary" and source_file not in seen_lm_files:
            lm_rows.extend(parse_lm_eval_ledger(row))

    summary_rows = build_summary_rows(ledger_rows, ppl_rows, lm_rows)
    missing_rows = build_missing_rows(summary_rows)

    workbook = load_or_create_workbook(template_path)
    write_sheet(workbook, "Summary", SUMMARY_COLUMNS, summary_rows)
    write_sheet(workbook, "PPL", PPL_COLUMNS, ppl_rows)
    write_sheet(workbook, "lm-eval", LM_EVAL_COLUMNS, lm_rows)
    write_sheet(workbook, "Standalone QKV", QKV_COLUMNS, qkv_rows)
    write_sheet(workbook, "Missing", MISSING_COLUMNS, missing_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print("=" * 80)
    print("Smart-Flip Workbook Export")
    print("=" * 80)
    print(f"Output:          {output_path}")
    print(f"Ledger rows:     {len(ledger_rows)}")
    print(f"PPL rows:        {len(ppl_rows)}")
    print(f"lm-eval rows:    {len(lm_rows)}")
    print(f"QKV rows:        {len(qkv_rows)}")
    print(f"Summary rows:    {len(summary_rows)}")
    print(f"Missing entries: {len(missing_rows)}")
    print(f"Generated at:    {datetime.now().isoformat(timespec='seconds')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
